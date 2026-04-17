# processing_engine.py
import asyncio
import concurrent.futures
import os
import time
import gc
from io import BytesIO
from datetime import datetime
import aiohttp
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from utils import (
    extract_urls, async_download_image_bytes, format_duration
)
from image_metrics import (
    pil_from_bytes, compute_sharpness_pil, detect_white_borders,
    check_first_photo_bg, analyze_text_content, detect_urls_from_text,
    detect_phone_numbers_from_text, detect_qr_codes, detect_watermark_advanced, analyze_and_classify_photo,
    detect_transparency_in_bytes, detect_1px_border, is_low_contrast_image
)

PAUSE_POLL_INTERVAL_SEC = 0.5


def find_id_column(df):
    """Знаходить колонку-ідентифікатор товару за пріоритетом.

    Спочатку точний збіг «ID», потім регістронезалежний пошук
    серед поширених назв (id, goods_id, код, артикул, product_id).
    Якщо нічого не знайдено — повертає першу колонку.
    """
    # Пріоритет 1: точний збіг
    if "ID" in df.columns:
        return "ID"
    # Пріоритет 2: регістронезалежний пошук
    id_names = {"id", "goods_id", "код", "артикул", "product_id"}
    for col in df.columns:
        if str(col).lower().strip() in id_names:
            return col
    # Запасний варіант: перша колонка
    return df.columns[0]


def _make_details_template(product_id, photo_index, url, options):
    """Створює базовий шаблон словника результатів для одного фото."""
    details = {
        "ID товару": product_id,
        "Фото (порядковий номер в КТ)": photo_index,
        "Посилання на фото": url,
        "Загальна оцінка якості фото": "Погане",
        "Опис виявлених недоліків": "",
        "Ширина": 0,
        "Висота": 0,
        "Різкість": 0.0,
        "Debug Info": "",
    }
    if options.get("check_shadows"):
        details["Тіні на головному фото"] = "Ні"
    if options.get("check_borders"):
        details["Некадровані фото"] = "Ні"
    if options.get("check_logos"):
        details["З логотипом"] = "Ні"
    if options.get("check_watermarks"):
        details["Водяний знак"] = "Ні"
    if options.get("check_rus_text"):
        details["Російський текст"] = "Ні"
    if options.get("check_qr_url"):
        details["Наявність URL або QR-коду на фото"] = "Ні"
    if options.get("check_phone_numbers"):
        details["Номери телефонів"] = "Ні"
    return details


def _make_preview_payload(image_data, max_side=460, quality=80):
    """Стискає зображення для GUI-превʼю, щоб не перевантажувати Tk."""
    img = pil_from_bytes(image_data)
    if img is None:
        return image_data
    try:
        img.thumbnail((max_side, max_side))
        buf = BytesIO()
        img.save(buf, format="JPEG", quality=quality, optimize=True)
        return buf.getvalue()
    except Exception:
        return image_data
    finally:
        img.close()


def photo_worker_sync(task_data, conf, data):
    """Синхронний аналіз зображення (CPU-навантаження).

    Призначений для запуску через loop.run_in_executor().
    Звільняє сирі байти (del data) одразу після створення PIL-образу,
    щоб уникнути Memory Leak при обробці 10 000+ фото.
    """
    product_id = task_data["product_id"]
    url = task_data["url"]
    photo_index = task_data["photo_index"]
    options = conf.get("options", {})

    details = _make_details_template(product_id, photo_index, url, options)

    metrics_results = {}

    # Перевірка прозорості до конвертації
    is_transp, tr_reason = detect_transparency_in_bytes(data)
    if is_transp:
        metrics_results["is_transparent"] = True
        metrics_results["transparency_reason"] = tr_reason

    # Конвертація + негайне звільнення байтів
    img = pil_from_bytes(data)
    del data  # Звільняємо сирі байти одразу

    if img is None:
        details["Опис виявлених недоліків"] = "Файл пошкоджено/Не фото"
        return details, f"[ID:{product_id}] ⚠️ Файл не є зображенням"

    try:
        width, height = img.size
        sharpness = compute_sharpness_pil(img)
        details["Різкість"] = round(sharpness, 2)
        details["Ширина"] = width
        details["Висота"] = height

        # Detect transparent/clear products (uniformly bright image).
        # Their sharpness is inherently near zero due to lack of contrast,
        # not due to blur — mark so the classifier can skip "Дуже розмите".
        if is_low_contrast_image(img):
            metrics_results["is_low_contrast_image"] = True

        important_log = []
        if is_transp:
            important_log.append("Прозорий фон")

        if options.get("check_shadows") and photo_index == 1:
            shadow_mode = int(conf.get("shadow_mode", 2))
            shadow_mode = min(4, max(1, shadow_mode))
            mode_profiles = conf.get("shadow_mode_profiles", {})
            profile = mode_profiles.get(str(shadow_mode), {}) if isinstance(mode_profiles, dict) else {}
            shadow_tolerance = int(profile.get("shadow_tolerance", conf.get("shadow_threshold", 50)))
            white_v_min = int(profile.get("white_v_min", 205))
            white_s_max = int(profile.get("white_s_max", 25))
            has_problem, reason = check_first_photo_bg(
                img,
                shadow_tolerance=shadow_tolerance,
                white_v_min=white_v_min,
                white_s_max=white_s_max,
            )
            metrics_results["has_shadows"] = has_problem
            metrics_results["shadows_reason"] = reason
            if has_problem:
                details["Тіні на головному фото"] = "Так"
                important_log.append(f"Тіні ({reason})")

        if options.get("check_borders"):
            border_ratio = conf.get("border_ratio", 0.1)
            has_borders, reason = detect_white_borders(img, border_ratio=border_ratio)
            metrics_results["has_white_borders"] = has_borders
            metrics_results["borders_reason"] = reason
            if has_borders:
                details["Некадровані фото"] = "Так"
                important_log.append(f"Поля ({reason})")

        if options.get("check_1px_border"):
            has_1px, reason_1px = detect_1px_border(img)
            metrics_results["has_1px_border"] = has_1px
            metrics_results["1px_border_reason"] = reason_1px
            if has_1px:
                important_log.append("Рамка 1px")

        ocr_wm_text = None
        word_count = 0
        qr_urls_found = []

        check_text_needed = (
            options.get("check_rus_text")
            or options.get("check_qr_url")
            or options.get("check_logos")
            or options.get("check_watermarks")
            or options.get("check_phone_numbers")
        )

        if check_text_needed:
            full_text, has_rus, is_rozetka_logo, ocr_wm_text, word_count = (
                analyze_text_content(img)
            )

            if options.get("check_logos") and is_rozetka_logo:
                metrics_results["has_logo"] = True
                details["З логотипом"] = "Так"
                important_log.append("Лого Rozetka")

            if options.get("check_rus_text") and has_rus:
                metrics_results["has_rus_text"] = True
                details["Російський текст"] = "Так"
                important_log.append("Рос. текст")

            if options.get("check_qr_url"):
                urls = detect_urls_from_text(full_text)
                if urls:
                    qr_urls_found.extend(urls)

                has_qr, qr_data = detect_qr_codes(img)
                if has_qr:
                    qr_urls_found.append(f"QR:{qr_data}")

                if qr_urls_found:
                    metrics_results["has_qr_url"] = True
                    metrics_results["qr_url_data"] = "; ".join(qr_urls_found)
                    details["Наявність URL або QR-коду на фото"] = "Так"
                    important_log.append(f"URL/QR ({len(qr_urls_found)})")

            if options.get("check_phone_numbers"):
                phones = detect_phone_numbers_from_text(full_text)
                if phones:
                    metrics_results["has_phone_numbers"] = True
                    metrics_results["phone_numbers_data"] = "; ".join(phones)
                    details["Номери телефонів"] = "Так"
                    important_log.append(f"Телефони ({len(phones)})")

        if options.get("check_watermarks"):
            has_wm, reason = detect_watermark_advanced(
                img, ocr_wm_text=ocr_wm_text, word_count=word_count
            )
            metrics_results["has_watermark"] = has_wm
            metrics_results["watermark_reason"] = reason
            if has_wm:
                details["Водяний знак"] = "Так"
                important_log.append("Watermark")

        status, reason_str, debug_str = analyze_and_classify_photo(
            width, height, sharpness, conf, metrics_results
        )
        details["Загальна оцінка якості фото"] = status
        details["Опис виявлених недоліків"] = reason_str

        if metrics_results.get("qr_url_data") and "QR" not in debug_str:
            debug_str += f"; Found: {metrics_results['qr_url_data']}"

        details["Debug Info"] = debug_str.lstrip("; ")

        log_msg = None
        if status == "Погане" or important_log:
            reasons_short = ", ".join(important_log) if important_log else reason_str
            log_msg = f"[ID:{product_id}] ❌ {reasons_short}"

        return details, log_msg

    finally:
        # Гарантовано закриваємо PIL-зображення для звільнення пам'яті
        img.close()

def regenerate_status_from_details(details_path):
    """
    Оновлює файл статусів на основі відредагованого файлу деталей.
    Використовує find_id_column для гнучкого визначення колонки ID.
    """
    try:
        base_dir = os.path.dirname(details_path)
        filename = os.path.basename(details_path)

        if "_Деталі" in filename:
            status_filename = filename.replace("_Деталі", "_Статус")
        else:
            return {"error": "Файл не містить '_Деталі'. Неможливо знайти пару."}

        status_path = os.path.join(base_dir, status_filename)

        if not os.path.exists(status_path):
            return {"error": f"Не знайдено файл статусів: {status_filename}"}

        df_details = pd.read_excel(details_path, engine="openpyxl", dtype=str)
        df_status = pd.read_excel(status_path, engine="openpyxl", dtype=str)

        # Гнучке визначення колонки ID
        details_id_col = find_id_column(df_details)
        status_id_col = find_id_column(df_status)

        # Нормалізуємо назви для групування
        df_details = df_details.rename(columns={details_id_col: "_ID_KEY"})
        df_status = df_status.rename(columns={status_id_col: "_ID_KEY"})

        details_grouped = df_details.groupby("_ID_KEY")

        new_statuses, new_problems, new_totals = [], [], []
        new_good, new_bad, new_mid = [], [], []

        for pid in df_status["_ID_KEY"]:
            if pid in details_grouped.groups:
                grp = details_grouped.get_group(pid)
                s_list = grp["Загальна оцінка якості фото"].tolist()

                if "Погане" in s_list:
                    fin_status = "Погане"
                elif "Середнє" in s_list:
                    fin_status = "Середнє"
                else:
                    fin_status = "Хороше"

                probs = []
                for _, r in grp.iterrows():
                    st = r.get("Загальна оцінка якості фото", "")
                    reason = r.get("Опис виявлених недоліків", "")
                    photo_idx = r.get("Фото (порядковий номер в КТ)", "?")
                    if st != "Хороше" and pd.notna(reason) and reason:
                        probs.append(f"{photo_idx} ({reason})")

                new_statuses.append(fin_status)
                new_problems.append("; ".join(probs))
                new_totals.append(len(s_list))
                new_good.append(s_list.count("Хороше"))
                new_bad.append(s_list.count("Погане"))
                new_mid.append(s_list.count("Середнє"))
            else:
                new_statuses.append("Хороше")
                new_problems.append("")
                new_totals.append(0)
                new_good.append(0)
                new_bad.append(0)
                new_mid.append(0)

        # Повертаємо оригінальну назву колонки
        df_status = df_status.rename(columns={"_ID_KEY": status_id_col})
        df_status["Загальна оцінка якості фото"] = new_statuses
        df_status["Проблемні фото"] = new_problems
        df_status["Всього фото"] = new_totals
        df_status["К-ть Хороших"] = new_good
        df_status["К-ть Поганих"] = new_bad
        df_status["К-ть Середніх"] = new_mid

        try:
            df_status.to_excel(status_path, index=False, engine="openpyxl")
            format_excel_header(status_path)
            return {"success": True, "path": status_path, "count": len(df_status)}
        except PermissionError:
            base, ext = os.path.splitext(status_path)
            fallback_path = f"{base}_UPDATED_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            df_status.to_excel(fallback_path, index=False, engine="openpyxl")
            format_excel_header(fallback_path)
            return {
                "success": True,
                "path": fallback_path,
                "count": len(df_status),
                "warning": "Основний файл зайнятий (відкритий в іншій програмі). Збережено копію з суфіксом _UPDATED."
            }

    except Exception as e:
        return {"error": f"Update Error: {str(e)}"}
        
def format_excel_header(path):
    try:
        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.row_dimensions[1].height = 50
            for cell in ws[1]:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                cell.font = Font(bold=True)
        wb.save(path)
    except Exception:
        pass


async def process_file(input_path, conf, gui_callback, manual_url_column, pause_event, stop_event):
    """Асинхронна обробка файлу Excel/CSV з фотографіями товарів.

    Використовує aiohttp для мережевих запитів та asyncio.Event для
    паузи/відновлення без зависання GUI.  CPU-навантаження (аналіз зображень)
    виконується в окремому пулі потоків (ThreadPoolExecutor).
    """
    def log(msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] {msg}")
        if gui_callback:
            try:
                gui_callback(f"[{ts}] {msg}")
            except Exception:
                pass

    # ---- Читання вхідного файлу ----
    try:
        ext = os.path.splitext(input_path)[1].lower()
        if ext == ".csv":
            df = pd.read_csv(input_path, dtype=str)
        else:
            df = pd.read_excel(input_path, engine="openpyxl", dtype=str)
    except Exception as e:
        return {"error": f"Open Error: {e}"}

    if df.shape[0] == 0:
        return {"error": "Файл порожній"}

    # ---- Визначення колонок ----
    id_col = find_id_column(df)

    best_col = None
    if manual_url_column and manual_url_column in df.columns:
        best_col = manual_url_column
    else:
        log("Авто-пошук колонки з URL...")
        for col in df.columns:
            if any(x in str(col).lower() for x in ["img", "url", "photo", "link", "foto"]):
                best_col = col
                break
        if not best_col:
            best_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]

    log(f"Колонка ID: '{id_col}' | Колонка URL: '{best_col}'")

    # ---- Побудова черги завдань із дедуплікацією ----
    all_tasks = []
    product_photo_map = {}
    url_first_seen = {}   # url -> True, перше входження по всьому файлу

    for idx, row in df.iterrows():
        product_id = str(row.get(id_col, "")).strip()
        if not product_id:
            product_id = f"ROW_{idx + 2}"

        cell = row.get(best_col, "")
        urls = extract_urls(cell)

        product_photo_map[product_id] = []
        seen_in_product = set()  # Дедуплікація всередині одного товару

        if not urls:
            all_tasks.append({
                "product_id": product_id, "url": "", "photo_index": 1,
                "is_empty": True, "is_duplicate": False, "cross_dup": False,
            })
            product_photo_map[product_id].append(1)
        else:
            for i, url in enumerate(urls[:100], start=1):
                product_photo_map[product_id].append(i)

                # Дубль всередині товару — той самий URL знову
                if url in seen_in_product:
                    all_tasks.append({
                        "product_id": product_id, "url": url, "photo_index": i,
                        "is_empty": False, "is_duplicate": True, "cross_dup": False,
                    })
                    continue

                seen_in_product.add(url)

                # Дубль по всьому файлу — URL зустрічається в іншому товарі
                if url in url_first_seen:
                    all_tasks.append({
                        "product_id": product_id, "url": url, "photo_index": i,
                        "is_empty": False, "is_duplicate": False, "cross_dup": True,
                    })
                else:
                    url_first_seen[url] = True
                    all_tasks.append({
                        "product_id": product_id, "url": url, "photo_index": i,
                        "is_empty": False, "is_duplicate": False, "cross_dup": False,
                    })

    total_photos = len(all_tasks)
    if gui_callback:
        gui_callback(("progress_max", total_photos))

    concurrency = int(conf.get("concurrency", 4))
    options = conf.get("options", {})

    # Розділяємо завдання на «оригінали» та «дублі»
    original_tasks = [t for t in all_tasks if not t["is_duplicate"] and not t["cross_dup"]]
    dup_tasks = [t for t in all_tasks if t["is_duplicate"] or t["cross_dup"]]

    all_results = []
    url_result_cache = {}  # url -> details dict (для cross_dup)
    stats = {"Good": 0, "Bad": 0, "Medium": 0, "Error": 0}
    start_time = time.time()
    processed_count = 0

    loop = asyncio.get_running_loop()
    cpu_executor = concurrent.futures.ThreadPoolExecutor(max_workers=concurrency)

    # ---- Фаза 1: Асинхронна обробка оригінальних URL ----
    headers = {
        "User-Agent": "PhotoQualityChecker/11.0",
        "Accept": "image/*, */*",
        "Accept-Encoding": "identity",
    }
    connector = aiohttp.TCPConnector(limit=concurrency * 2)

    async with aiohttp.ClientSession(headers=headers, connector=connector) as session:
        semaphore = asyncio.Semaphore(concurrency)

        async def wait_until_resumed_or_stopped():
            while not stop_event.is_set():
                if pause_event.is_set():
                    return True
                try:
                    await asyncio.wait_for(pause_event.wait(), timeout=PAUSE_POLL_INTERVAL_SEC)
                except asyncio.TimeoutError:
                    continue
            return False

        async def process_original(task):
            # Чекаємо на відновлення, якщо стоїть пауза
            if not await wait_until_resumed_or_stopped():
                return None

            product_id = task["product_id"]
            photo_index = task["photo_index"]
            url = task["url"]

            base = _make_details_template(product_id, photo_index, url, options)

            if task["is_empty"]:
                base["Опис виявлених недоліків"] = "Немає фото"
                return base, None

            # Завантаження зображення (async I/O)
            data, err = await async_download_image_bytes(url, session, semaphore)

            if data is None:
                base["Опис виявлених недоліків"] = "Не вдалося завантажити"
                base["Debug Info"] = str(err)
                return base, f"[ID:{product_id}] ⚠️ Помилка завантаження: {err}"

            # Надсилаємо перше фото кожного товару як превʼю в GUI
            if photo_index == 1 and gui_callback:
                gui_callback(("preview_image", _make_preview_payload(data)))

            # Аналіз зображення (CPU-bound, у пулі потоків)
            try:
                result = await loop.run_in_executor(
                    cpu_executor, photo_worker_sync, task, conf, data
                )
                return result
            except Exception as e:
                base["Опис виявлених недоліків"] = f"Помилка аналізу: {e}"
                return base, f"[ID:{product_id}] ❌ Критична помилка: {e}"

        pending_tasks = set()
        task_index = 0
        originals_count = len(original_tasks)

        while (task_index < originals_count or pending_tasks) and not stop_event.is_set():
            while task_index < originals_count and len(pending_tasks) < concurrency and not stop_event.is_set():
                if not await wait_until_resumed_or_stopped():
                    break
                pending_tasks.add(asyncio.create_task(process_original(original_tasks[task_index])))
                task_index += 1

            if not pending_tasks:
                break

            done, pending_tasks = await asyncio.wait(
                pending_tasks, return_when=asyncio.FIRST_COMPLETED
            )

            for done_task in done:
                result = done_task.result()
                processed_count += 1

                elapsed = time.time() - start_time
                avg = elapsed / processed_count if processed_count > 0 else 0
                # ETA враховує всі завдання (включно з дублями)
                remaining = total_photos - processed_count
                eta_seconds = avg * remaining

                if result is not None:
                    details, log_msg = result if isinstance(result, tuple) else (result, None)
                    if details:
                        all_results.append(details)
                        # Кешуємо для cross_dup-завдань
                        url = details.get("Посилання на фото", "")
                        if url:
                            url_result_cache[url] = details

                        st = details.get("Загальна оцінка якості фото", "Error")
                        if st == "Хороше":
                            stats["Good"] += 1
                        elif st == "Погане":
                            stats["Bad"] += 1
                        elif st == "Середнє":
                            stats["Medium"] += 1
                        else:
                            stats["Error"] += 1

                        if log_msg:
                            log(log_msg)

                if processed_count % 50 == 0:
                    log(f"--- {processed_count}/{total_photos} ---")
                    gc.collect()

                if gui_callback:
                    gui_callback(("progress_update", (processed_count, total_photos, eta_seconds)))

                if stop_event.is_set():
                    break

        if stop_event.is_set() and pending_tasks:
            for p in pending_tasks:
                p.cancel()
            await asyncio.gather(*pending_tasks, return_exceptions=True)

    cpu_executor.shutdown(wait=False)

    # ---- Фаза 2: Обробка дублів (синхронно, без I/O) ----
    for task in dup_tasks:
        product_id = task["product_id"]
        photo_index = task["photo_index"]
        url = task["url"]
        base = _make_details_template(product_id, photo_index, url, options)

        if task["is_duplicate"]:
            # Дубль всередині товару
            base["Опис виявлених недоліків"] = "Дубль"
            all_results.append(base)
        elif task["cross_dup"]:
            # Дубль посилання з іншого товару — копіюємо результат
            cached = url_result_cache.get(url)
            if cached:
                dup_details = cached.copy()
                dup_details["ID товару"] = product_id
                dup_details["Фото (порядковий номер в КТ)"] = photo_index
                all_results.append(dup_details)
                st = dup_details.get("Загальна оцінка якості фото", "Error")
                if st == "Хороше":
                    stats["Good"] += 1
                elif st == "Погане":
                    stats["Bad"] += 1
                elif st == "Середнє":
                    stats["Medium"] += 1
            else:
                base["Опис виявлених недоліків"] = "Не вдалося отримати результат (дубль)"
                all_results.append(base)
                stats["Error"] += 1

        processed_count += 1
        if gui_callback:
            elapsed = time.time() - start_time
            avg = elapsed / max(processed_count, 1)
            remaining = total_photos - processed_count
            gui_callback(("progress_update", (processed_count, total_photos, avg * remaining)))

    if not all_results:
        return {"error": "No results."}

    # ---- Збереження результатів ----
    log("Збереження...")
    try:
        details_df = pd.DataFrame(all_results)
        details_df["ID товару"] = pd.Categorical(
            details_df["ID товару"], categories=product_photo_map.keys(), ordered=True
        )
        details_df = details_df.sort_values(by=["ID товару", "Фото (порядковий номер в КТ)"])

        OPTIONAL_COLS_MAP = {
            "check_shadows": "Тіні на головному фото",
            "check_borders": "Некадровані фото",
            "check_logos": "З логотипом",
            "check_watermarks": "Водяний знак",
            "check_rus_text": "Російський текст",
            "check_qr_url": "Наявність URL або QR-коду на фото",
            "check_phone_numbers": "Номери телефонів",
        }
        cols = ["ID товару", "Фото (порядковий номер в КТ)", "Посилання на фото", "Загальна оцінка якості фото", "Опис виявлених недоліків"]
        for key, col_name in OPTIONAL_COLS_MAP.items():
            if options.get(key) and col_name in details_df.columns:
                has_yes = details_df[col_name].fillna("").astype(str).str.strip().eq("Так").any()
                if has_yes:
                    cols.append(col_name)
        cols.extend(["Ширина", "Висота", "Різкість", "Debug Info"])

        for c in cols:
            if c not in details_df.columns:
                details_df[c] = ""
        details_df = details_df[cols]

        df_out = df.copy()
        status_map, problems_map = {}, {}
        stats_data = []

        for pid, grp in details_df.groupby("ID товару", observed=True):
            s_list = grp["Загальна оцінка якості фото"].tolist()
            fin = "Хороше"
            if "Погане" in s_list:
                fin = "Погане"
            elif "Середнє" in s_list:
                fin = "Середнє"
            status_map[pid] = fin
            probs = [
                f"{r['Фото (порядковий номер в КТ)']} ({r['Опис виявлених недоліків']})"
                for _, r in grp.iterrows()
                if r["Загальна оцінка якості фото"] != "Хороше" and r["Опис виявлених недоліків"]
            ]
            problems_map[pid] = "; ".join(probs)
            stats_data.append({
                "ID товару": pid,
                "Всього фото": len(s_list),
                "К-ть Хороших": s_list.count("Хороше"),
                "К-ть Поганих": s_list.count("Погане"),
                "К-ть Середніх": s_list.count("Середнє"),
            })

        stats_df = pd.DataFrame(stats_data).set_index("ID товару") if stats_data else pd.DataFrame()
        df_out["Загальна оцінка якості фото"] = df_out[id_col].map(status_map).fillna("N/A")
        df_out["Проблемні фото"] = df_out[id_col].map(problems_map).fillna("")

        if not stats_df.empty:
            for c in ["Всього фото", "К-ть Хороших", "К-ть Поганих", "К-ть Середніх"]:
                df_out[c] = df_out[id_col].map(stats_df[c]).fillna(0)

        base_dir = os.path.dirname(input_path)
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        suffix = "_PARTIAL" if stop_event.is_set() else ""
        p1 = os.path.join(base_dir, f"{base_name}_Деталі{suffix}.xlsx")
        p2 = os.path.join(base_dir, f"{base_name}_Статус{suffix}.xlsx")

        details_df.to_excel(p1, index=False, engine="openpyxl")
        df_out.to_excel(p2, index=False, engine="openpyxl")
        format_excel_header(p1)
        format_excel_header(p2)

        elapsed = time.time() - start_time
        elapsed_str = format_duration(elapsed)

        log("=" * 30)
        log(f"🏁 DONE in {elapsed_str}")
        log(f"✅ Ok: {stats['Good']} | ❌ Bad: {stats['Bad']} | ⚠️ Mid: {stats['Medium']}")
        log("=" * 30)

        res = {"details": p1, "with_status": p2}
        if stop_event.is_set():
            res["stopped"] = True
        return res

    except Exception as e:
        return {"error": f"Save Error: {e}"}
